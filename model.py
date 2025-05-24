import logging
import pandas as pd
import os
import pandas as pd
from threading import Thread
from concurrent.futures import ThreadPoolExecutor, as_completed
import pyarrow.csv as pv
import hashlib
import pickle
import requests
from police_forces import police_forces, offence_types
import re
from math import radians, sin, cos, sqrt, atan2, isnan
import numpy as np
import scipy.stats as stats
from datetime import datetime
from multiprocessing import Pool, cpu_count
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Color, PatternFill, Alignment, Fill
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.utils import get_column_letter
from final_score_weightings import final_weightings as fsw
from concurrent.futures import ProcessPoolExecutor
from weightings_dictionary import weightings_dict
import multiprocessing

from scipy import stats
from functools import lru_cache
from tqdm import tqdm
crime_cache = {}


# Set up logging
logging.basicConfig(level=logging.INFO,
                    format='%(asctime)s - %(levelname)s - %(message)s')

crime_categories = ["Anti-Social Behaviour", "Burglary", "Robbery", "Vehicle Crimes", "Violent Crimes", "Shoplifting",
                    "Criminal Damage & Arson", "Other Theft", "Drugs", "Other Crimes", "Bike Theft", "Weapons", "Order", "Theft From Person"]


def parallel_process_data(func, df, data_sources=None, args=()):
    """
    Generic parallel processing function that handles both single and multiple data sources
    """
    with ThreadPoolExecutor(max_workers=min(32, multiprocessing.cpu_count())) as executor:
        if data_sources is None:
            # For functions that only need the row
            results = list(executor.map(
                lambda row: func(row[1]),
                df.iterrows()
            ))
        elif isinstance(data_sources, tuple):
            # For functions expecting multiple data sources as positional arguments
            results = list(executor.map(
                lambda row: func(row[1], *data_sources, *args),
                df.iterrows()
            ))
        elif isinstance(data_sources, dict):
            # For functions expecting data sources as keyword arguments
            if args:
                results = list(executor.map(
                    lambda row: func(row[1], *args, **data_sources),
                    df.iterrows()
                ))
            else:
                results = list(executor.map(
                    lambda row: func(row[1], **data_sources),
                    df.iterrows()
                ))
        else:
            # For functions expecting a single data source
            results = list(executor.map(
                lambda row: func(row[1], data_sources, *args),
                df.iterrows()
            ))
    return zip(*results) if isinstance(results[0], tuple) else results




def save_model(df, user, location):
    now = datetime.now().strftime("%Y-%m-%d %H-%M-%S")
    excel_file = rf"C:\Users\{user}\{location}\Risk - Documents\MODELS\Created Templates\Model {now}.xlsx"
    df.to_excel(excel_file, index=False)

    wb = load_workbook(excel_file)
    ws = wb.active

    # Define border, fill, and alignment styles
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    title_fill = PatternFill(start_color='009FE3', end_color='009FE3', fill_type='solid')  # Yellow fill
    center_aligned_text = Alignment(horizontal='center', vertical='center')

    # Apply styling and adjust column widths
    cols = df.columns.to_list()
    for i, col in enumerate(cols, 1):  # Start enumeration at 1 to match Excel's 1-based index
        column_letter = get_column_letter(i)
        max_length = max(len(str(value)) for value in df[col])  # Find maximum length of column values
        max_length = max(max_length, len(str(col)))  # Compare with header length
        ws.column_dimensions[column_letter].width = max_length + 2  # Set the width based on max length + some padding

        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=i)
            cell.border = thin_border
            cell.alignment = center_aligned_text
            if row == 1:
                cell.fill = title_fill  # Apply fill only to the first row

        # Apply conditional formatting
        cell_range = f"{column_letter}2:{column_letter}{ws.max_row}"
        if col not in ['Risk Output', 'postcode', 'corrected_postcode']:
            ws.conditional_formatting.add(
                cell_range, 
                ColorScaleRule(
                    start_type='num', start_value=1, start_color='FF00803E',
                    mid_type='num', mid_value=3, mid_color='FFFCB813',
                    end_type='num', end_value=5, end_color='FFC42B1C'))
        else:
            ws.conditional_formatting.add(
                cell_range, 
                ColorScaleRule(
                    start_type='num', start_value=20, start_color='FF00803E',
                    mid_type='num', mid_value=60, mid_color='FFFCB813',
                    end_type='num', end_value=100, end_color='FFC42B1C'))

    wb.save(excel_file) 

def get_column_by_name(ws, column_name):
    for idx, col in enumerate(ws[1], 1):  # First row
        if col.value == column_name:
            return get_column_letter(idx)
    return None

def get_scottish_council_area(row, data):
    if row['country'] == 'Scotland':
       sector = row['sector']
       try:
           value = data[data['PostcodeSector'] == sector]['CouncilArea2019Code'].values[0]
        #    print(value)
           return value
       except:
           return None
    else:
        return None

def get_pubs_data(row, pubs_data):
    lat = row['latitude']
    lon = row['longitude']

    distances = pubs_data.apply(lambda row: haversine_distance(
        lat, lon, row['lat'], row['long']), axis=1)
    within_radius = pubs_data[distances <= 1.5]
    rank = rank_pubs_data(len(within_radius))
    return len(within_radius), rank


def rank_pubs_data(count):
    if count <= 0:
        return 1
    elif count <= 3:
        return 2
    elif count <= 6:
        return 3
    elif count <= 9:
        return 4
    elif count <= 20:
        return 5
    elif count <= 30:
        return 6
    elif count <= 50:
        return 7
    elif count <= 75:
        return 8
    elif count <= 100:
        return 9
    else:
        return 10


def get_population_density_data(row, uk_population_density_data, ni_population_density_data, scotland_population_density_data):
    admin_district = row['admin_district']
    country = row['country']

    if country == 'England' or country == 'Wales':
        try:
            data = uk_population_density_data[uk_population_density_data['admin_district']
                                              == admin_district]['population_density'].values[0]
            rank = assign_score(stats.percentileofscore(
                uk_population_density_data['population_density'].values, data))
        except:
            data = None
            rank = None
    if country == 'Scotland':
        try:
            data = scotland_population_density_data[scotland_population_density_data['admin_district']
                                                    == admin_district]['population_density'].values[0]
            rank = assign_score(stats.percentileofscore(
                scotland_population_density_data['population_density'].values, data))
        except:
            data = None
            rank = None
    if country == 'Northern Ireland':
        try:
            data = ni_population_density_data[ni_population_density_data['admin_district']
                                              == admin_district]['population_density'].values[0]
            rank = assign_score(stats.percentileofscore(
                ni_population_density_data['population_density'].values, data))
        except:
            data = None
            rank = None

    return data, rank

def get_homelessness_data(row, uk_homelessness, ni_homelessness, scotland_homelessness):
    admin_district = row['admin_district']
    country = row['country']

    if country == 'England' or country == 'Wales':
        try:
            data = uk_homelessness[uk_homelessness['admin_district']
                                              == admin_district]['value'].values[0]
            rank = assign_score(stats.percentileofscore(
                uk_homelessness['value'].values, data))
        except:
            data = None
            rank = None
    if country == 'Scotland':
        try:
            data = scotland_homelessness[scotland_homelessness['admin_district']
                                                    == admin_district]['value'].values[0]
            rank = assign_score(stats.percentileofscore(
                scotland_homelessness['value'].values, data))
        except:
            data = None
            rank = None
    if country == 'Northern Ireland':
        try:
            data = ni_homelessness[ni_homelessness['admin_district']
                                              == admin_district]['value'].values[0]
            rank = assign_score(stats.percentileofscore(
                ni_homelessness['value'].values, data))
        except:
            data = None
            rank = None

    return data, rank

def assign_risk_scores(data):
    # Initialize with postcode
    result = {'corrected_postcode': data['corrected_postcode']}
    
    # Initialize all risks with None by default
    for risk in weightings_dict.keys():
        result[risk] = None
    
    for risk, value in weightings_dict.items():
        # print(f"\nProcessing {risk} for {data['corrected_postcode']}")
        metrics = value.copy()
        counted_metrics = value.copy()
        metrics_to_remove = []

        # Check each metric
        for metric, weighting in value.items():
            metric_key = f'{metric.lower()}_rank'
            # Check the actual value before deciding to remove
            current_value = data.get(metric_key)
            
            should_remove = (
                weighting == 0 or 
                current_value is None or 
                pd.isna(current_value) or 
                current_value == 0  # Add this if 0 should be treated as missing
            )
            
            if should_remove:
                metrics_to_remove.append(metric)
                # print(f"Removing {metric} because value is {current_value}")

        # Remove identified metrics
        for metric in metrics_to_remove:
            if metric in counted_metrics:
                counted_metrics.pop(metric)
                
        # print(f"Metrics remaining: {counted_metrics}")
        risk_score_counter = 0
        # print(risk, "-", counted_metrics)
        
        if counted_metrics:
            # Normalize weights to sum to 1
            total_weight = sum(counted_metrics.values())
            if total_weight > 0:
                multiplier = 1 / total_weight
                counted_metrics = {metric: value * multiplier 
                                 for metric, value in counted_metrics.items()}
            
                # Calculate risk score
                for metric, value in counted_metrics.items():
                    metric_key = f'{metric.lower()}_rank'
                    rank_value = data[metric_key]
                    if pd.notna(rank_value) and rank_value != 0:
                        risk_score_counter += (rank_value * value)
                        # print(f"Added {metric}: {rank_value} * {value} = {rank_value * value}")
        
            # print(f"Final score for {risk}: {risk_score_counter}")
            if risk_score_counter > 0:
                result[risk] = round(risk_score_counter, 0)
            # If risk_score_counter is 0, the None value set at initialization will remain

    return pd.Series(result)

def get_student_data(row, uk_student_data, ni_student_data, scotland_student_data):
    admin_district = row['admin_district']
    country = row['country']

    if country == 'England' or country == 'Wales':
        try:
            data = uk_student_data[uk_student_data['admin_district']
                                   == admin_district]['student_population'].values[0]
            rank = assign_score(stats.percentileofscore(
                uk_student_data['student_population'].values, data))

        except:
            data = None
            rank = None
    if country == 'Northern Ireland':
        try:
            data = ni_student_data[ni_student_data['admin_district']
                                   == admin_district]['student_population'].values[0]
            rank = assign_score(stats.percentileofscore(
                ni_student_data['student_population'].values, data))

        except:
            data = None
            rank = None
    if country == 'Scotland':
        try:
            data = scotland_student_data[scotland_student_data['admin_district']
                                         == admin_district]['student_population'].values[0]
            rank = assign_score(stats.percentileofscore(
                scotland_student_data['student_population'].values, data))

        except:
            data = None
            rank = None

    return data, rank


def get_unemployment_data(row, unemployment_data, ni_unemployment_data, scotland_unemployment_data):
    admin_district = row['admin_district']

    if row['country'] == 'England' or row['country'] == 'Wales':
        try:
            data = unemployment_data[unemployment_data['admin_district']
                                     == admin_district]['unemployment'].values[0]
            rank = assign_score(stats.percentileofscore(
                unemployment_data['unemployment'].values, data))

        except:
            data = None
            rank = None

        return data, rank
    elif row['country'] == 'Northern Ireland':
        try:
            data = ni_unemployment_data[ni_unemployment_data['admin_district']
                                        == admin_district]['unemployment'].values[0]
            rank = assign_score(stats.percentileofscore(
                ni_unemployment_data['unemployment'].values, data))

        except:
            data = None
            rank = None

        return data, rank

    elif row['country'] == 'Scotland':
        try:
            data = scotland_unemployment_data[scotland_unemployment_data['admin_district']
                                              == admin_district]['unemployment'].values[0]
            rank = assign_score(stats.percentileofscore(
                scotland_unemployment_data['unemployment'].values, data))

        except:
            data = None
            rank = None

        return data, rank


def get_wellbeing_data(row, well_being_data, ni_wellbeing_data):
    admin_district = row['admin_district']
    country = row['country']

    if country != 'Northern Ireland':
        try:
            data = well_being_data[well_being_data['admin_district']
                                   == admin_district]['wellbeing'].values[0]
            rank = assign_score(stats.percentileofscore(
                well_being_data['wellbeing'].values, data))
        except:
            data = None
            rank = None

        return data, 6 - rank if rank is not None else None
    else:
        try:
            data = ni_wellbeing_data[ni_wellbeing_data['admin_district']
                                     == admin_district]['wellbeing'].values[0]
            rank = assign_score(stats.percentileofscore(
                ni_wellbeing_data['wellbeing'].values, data))
        except:
            data = None
            rank = None

        return data, 6 - rank if rank is not None else None


def get_transport_links_data(row, transport_links_data, radius):
    lat = row['latitude']
    long = row['longitude']

    distances = transport_links_data.apply(lambda row: haversine_distance(
        lat, long, row['lat'], row['long']), axis=1)
    within_radius = transport_links_data[distances <= radius]
    rank = rank_transport_links_data(len(within_radius))
    return len(within_radius), rank


def rank_transport_links_data(value):
    if value <= 0:
        return 1
    elif value <= 1:
        return 3
    elif value <= 2:
        return 4
    elif value <= 3:
        return 5
    elif value <= 4:
        return 6
    elif value <= 5:
        return 7
    elif value <= 6:
        return 8
    elif value <= 7:
        return 9
    else:
        return 10

def get_junctions_data(row, junctions_data, radius):
    lat = row['latitude']
    long = row['longitude']

    distances = junctions_data.apply(lambda row: haversine_distance(
        lat, long, row['Lat'], row['Long']), axis=1)
    within_radius = junctions_data[distances <= radius]
    rank = rank_junctions_data(len(within_radius))
    return len(within_radius), rank


def rank_junctions_data(value):
    if value <= 0:
        return 1
    elif value <= 1:
        return 3
    elif value <= 2:
        return 4
    elif value <= 3:
        return 5
    elif value <= 4:
        return 6
    elif value <= 5:
        return 7
    elif value <= 6:
        return 8
    elif value <= 7:
        return 9
    else:
        return 10


def get_schools_data(row, school_data, radius):
    lat = row['latitude']
    long = row['longitude']

    distances = school_data.apply(lambda row: haversine_distance(
        lat, long, row['latitude'], row['longitude']), axis=1)
    within_radius = school_data[distances <= radius]
    rank = rank_schools_data(len(within_radius))
    return len(within_radius), rank


def rank_schools_data(value):
    if value <= 0:
        return 1
    elif value <= 1:
        return 3
    elif value <= 2:
        return 4
    elif value <= 3:
        return 5
    elif value <= 4:
        return 6
    elif value <= 5:
        return 7
    elif value <= 6:
        return 8
    elif value <= 7:
        return 9
    else:
        return 10


def get_retail_park_data(row, park_data, radius):
    lat = row['latitude']
    long = row['longitude']

    distances = park_data.apply(lambda row: haversine_distance(
        lat, long, row['Lat'], row['Long']), axis=1)
    within_radius = park_data[distances <= radius]
    rank = rank_retail_park_data(len(within_radius))
    return len(within_radius), rank


def rank_retail_park_data(value):
    if value <= 0:
        return 1
    elif value <= 1:
        return 5
    elif value <= 2:
        return 6
    elif value <= 3:
        return 7
    elif value <= 4:
        return 8
    elif value <= 5:
        return 9
    else:
        return 10


def process_multiple_offences(outputs_df, offence_types, data_library):
    print('processing multiple offences...\n\n\n\n')
    # Start with the existing outputs_df
    result_df = outputs_df.copy()
    

    for offence in offence_types:
        temp_df = outputs_df.apply(
            lambda row: pd.Series(
                process_crime_data(
                    offence, 
                    data_library['uk_historic_detailed_police_reporting'],
                    data_library['scotland_historic_detailed_police_reporting'],
                    row  # now passing the entire row
                ),
                index=[f'{offence}_historic_count', f'{offence}_rank']
            ),
            axis=1  # specify axis=1 to apply function to each row
        )

        # Add the new columns to result_df
        result_df = pd.concat([result_df, temp_df], axis=1)

    return result_df

def process_chunk_historic(args):
    chunk_df, uk_hist_crime, scotland_hist_crime = args
    return chunk_df.apply(
        lambda row: process_historic_crime_data(row, uk_hist_crime, scotland_hist_crime),
        axis=1
    )

def process_all_historic_crime_data(outputs_df, uk_hist_crime, scotland_hist_crime):
    """Process all historic crime data in parallel with progress bar"""
    num_processes = max(1, cpu_count() - 1)
    total_rows = len(outputs_df)
    
    # For very small datasets, just process directly
    if total_rows < num_processes:
        print(f"\nProcessing {total_rows} rows directly (dataset too small for parallel processing)...")
        results = outputs_df.apply(
            lambda row: process_historic_crime_data(row, uk_hist_crime, scotland_hist_crime),
            axis=1
        )
        return pd.DataFrame(results)
    
    print(f"\nProcessing {total_rows} rows using {num_processes} processes for historic crime data...")
    
    # Split DataFrame into chunks, ensuring chunk_size is at least 1
    chunk_size = max(1, len(outputs_df) // num_processes)
    chunks = [outputs_df.iloc[i:i + chunk_size] for i in range(0, len(outputs_df), chunk_size)]
    
    # Prepare arguments for each chunk
    chunk_args = [(chunk, uk_hist_crime, scotland_hist_crime) for chunk in chunks]
    
    # Process chunks in parallel with progress bar
    with Pool(num_processes) as pool:
        results = list(tqdm(
            pool.imap(process_chunk_historic, chunk_args),
            total=len(chunks),
            desc="Processing historic crime data"
        ))
    
    print("Combining results...")
    final_results = pd.concat(results)
    print(f"Completed processing {total_rows} rows for historic crime data")
    return final_results

def process_historic_crime_data(row, uk_hist_crime, scotland_hist_crime):
    results = {'corrected_postcode': row['corrected_postcode']}
    if row['country'] in ['England', 'Wales']:
        for offence in offence_types:
            all_data = uk_hist_crime[uk_hist_crime['Offence Description'] == offence]
            row_data = all_data[all_data['Force Name'] == police_forces[row['police_region']]]
            if not row_data.empty:
                value = row_data['Number of Offences'].iloc[0]
                percentile = stats.percentileofscore(all_data['Number of Offences'], value)
                score = assign_score(percentile)
                results[f'{offence}_count'] = value
                results[f'{offence}_rank'] = score if value != 0 else 1
    elif row['country'] == 'Scotland':
        for offence in offence_types:
            all_data = scotland_hist_crime[scotland_hist_crime['uk_crime_classifications'] == offence]
            row_data = all_data[all_data['area_code'] == row['council_area']]
            if not row_data.empty:
                value = row_data['NUMBER_OF_DETECTED_CRIMES'].iloc[0]
                percentile = stats.percentileofscore(all_data['NUMBER_OF_DETECTED_CRIMES'], value)
                score = assign_score(percentile)
                results[f'{offence}_count'] = value
                results[f'{offence}_rank'] = score if value != 0 else 1
            else:
                results[f'{offence}_count'] = None
                results[f'{offence}_rank'] = None
    else:
        for offence in offence_types:
            results[f'{offence}_count'] = None
            results[f'{offence}_rank'] = None

    return pd.Series(results)
                

            

def process_crime_data(offence, crime_data, scottish_crime_data, row, cache={}):
    # Create cache key
    cache_key = (offence, row['country'], row.get('police_region', '') or row.get('council_area', ''))
    
    # Check cache first
    if cache_key in cache:
        return cache[cache_key]
    
    if row['country'] in ['England', 'Wales']:
        police_region_name = police_forces[row['police_region']]
        
        # Use boolean indexing once
        mask = (crime_data['Force Name'] == police_region_name) & (crime_data['Offence Description'] == offence)
        data = crime_data[mask]
        
        if data.empty:
            cache[cache_key] = (None, None)
            return None, None
            
        value = data['Number of Offences'].iloc[0]
        
        # Get pre-calculated percentile ranks if not already cached
        percentile_key = f'percentile_{offence}'
        if percentile_key not in cache:
            offence_mask = crime_data['Offence Description'] == offence
            all_offence_data = crime_data[offence_mask]['Number of Offences'].values
            cache[percentile_key] = all_offence_data
        
        percentile_rank = stats.percentileofscore(cache[percentile_key], value)
        score = assign_score(percentile_rank)
        
        cache[cache_key] = (value, score)
        return value, score
        
    elif row['country'] == 'Scotland':
        council_area = str(row['council_area']).strip()
        
        # Pre-calculate grouped data if not already cached
        if 'scottish_grouped' not in cache:
            cache['scottish_grouped'] = scottish_crime_data.groupby(
                ['area_code', 'uk_crime_classifications']
            )['NUMBER_OF_DETECTED_CRIMES'].sum().reset_index()
        
        scottish_crime_adjusted_data = cache['scottish_grouped']
        
        # Use boolean indexing once
        mask = (scottish_crime_adjusted_data['area_code'] == council_area) & \
               (scottish_crime_adjusted_data['uk_crime_classifications'] == offence)
        data = scottish_crime_adjusted_data[mask]
        
        try:
            value = data['NUMBER_OF_DETECTED_CRIMES'].sum()
        except:
            value = 0
            
        # Get pre-calculated percentile data if not already cached
        scot_percentile_key = f'scot_percentile_{offence}'
        if scot_percentile_key not in cache:
            offence_mask = scottish_crime_adjusted_data['uk_crime_classifications'] == offence
            all_offence_data = scottish_crime_adjusted_data[offence_mask]['NUMBER_OF_DETECTED_CRIMES'].values
            cache[scot_percentile_key] = all_offence_data
            
        percentile_rank = stats.percentileofscore(cache[scot_percentile_key], value)
        score = assign_score(percentile_rank)
        
        cache[cache_key] = (value, score)
        return value, score       


def haversine_distance(lat1, lon1, lat2, lon2):
    R = 6371  # Earth's radius in kilometers

    lat1, lon1, lat2, lon2 = map(radians, [lat1, lon1, lat2, lon2])
    dlat = lat2 - lat1
    dlon = lon2 - lon1

    a = sin(dlat/2)**2 + cos(lat1) * cos(lat2) * sin(dlon/2)**2
    c = 2 * atan2(sqrt(a), sqrt(1-a))
    distance = R * c

    return distance


def count_population_within_radius(row, df, radius_km, monthly_crime_data, ni_crime_data):
    # print('ni_crime_data', ni_crime_data)
    center_lat = row['latitude']
    center_lon = row['longitude']

    country = row['country']

    if country == 'England' or country == 'Wales':
        radius_km = radius_km * 1.60934

        within_radius = df[df.apply(lambda row: haversine_distance(center_lat, center_lon,
                                                                   row['Latitude'], row['Longitude']) <= radius_km, axis=1)]

        count = len(within_radius)
        postcodes = pd.DataFrame()
        postcodes['sector'] = within_radius['Postcode'].apply(
            lambda pc: get_postcode_breakdowns(pc)[0])
        sectors = postcodes['sector'].unique().tolist()

        crimes_df = pd.DataFrame(
            {'corrected_postcode': [row['corrected_postcode']]})

        # print(crimes_df)

        for crime in crime_categories:
            observable_crimes_df = monthly_crime_data[monthly_crime_data['Postcode Sector'].isin(
                sectors)]
            crime_count = observable_crimes_df[crime].sum() / len(sectors)

            # Get all crime counts for this crime type
            all_crime_counts = monthly_crime_data[crime]

            # Calculate the percentile rank of crime_count
            percentile = stats.percentileofscore(all_crime_counts, crime_count)
            score = assign_score(percentile)

            crimes_df[f'{crime} monthly_count'] = [crime_count]
            crimes_df[f'{crime} monthly_rank'] = [score]

        return crimes_df

    elif country == 'Northern Ireland':
        crimes_df = pd.DataFrame(
            {'corrected_postcode': [row['corrected_postcode']]})

        # print(crimes_df)

        df = ni_crime_data

        applicable_row = df[df['admin_district'] == row['admin_district']]
        # print(applicable_row)

        for crime in crime_categories:
            crime_count = applicable_row[crime][0]

            if crime_count is not None:
                percentile = stats.percentileofscore(df[crime], crime_count)
                score = assign_score(percentile)

            else:
                score = None

            crimes_df[f'{crime}_count'] = [crime_count]
            crimes_df[f'{crime} monthly_rank'] = [score]
            # print(row['corrected_postcode'], crime, crime_count, score)

        return crimes_df

    # print(postcodes)


def assign_score(percentile):
    if np.isnan(percentile) or percentile is None:
        return None
    elif percentile <= 10:
        return 1
    elif percentile <= 20:
        return 2
    elif percentile <= 30:
        return 3
    elif percentile <= 40:
        return 4
    elif percentile <= 50:
        return 5
    elif percentile <= 60:
        return 6
    elif percentile <= 70:
        return 7
    elif percentile <= 80:
        return 8
    elif percentile <= 90:
        return 9
    else:
        return 10

def process_internal_data(field, row, column):
    value = row[field]
    if value is not None and not pd.isna(value):
        percentile = stats.percentileofscore(column, value)
        score = assign_score(percentile)
        return score
    
def get_postcode_breakdowns(postcode):
    search = re.findall(
        r"^(((([A-Z][A-Z]{0,1})[0-9][A-Z0-9]{0,1}) {0,}[0-9])[A-Z]{2})$", postcode)
    # print(postcode)
    # print(search)
    try:
        sector, district, area = search[0][1], search[0][2], search[0][3]
        # print(sector, " | ", district, " | ", area)
        return sector, district, area
    except Exception as e:
        # print(e)
        return "", "", ""


def get_country(msoa):
    country_code = msoa[0]
    match country_code:
        case 'E':
            return 'England'
        case 'N':
            return 'Northern Ireland'
        case 'W':
            return 'Wales'
        case 'S':
            return 'Scotland'
        case _:
            return 'Unknown'


def correct_postcode(postcode):
    count = 1
    max_retries = 4  # Define a maximum number of retries to prevent an infinite loop

    # Retry logic starts here
    while count <= max_retries:
        url = f'https://api.postcodes.io/postcodes/{postcode}/autocomplete/'
        response = requests.get(url)

        if response.status_code == 200:
            data = response.json()
            if data['result']:
                # print(data['result'])
                corrected_pc = data['result'][0]
                return corrected_pc
                break  # Success, exit the retry loop
            else:
                # print(f'retrying {postcode}')
                postcode = postcode[:count*-1]
        else:
            # print(f"Failed to get a successful response, status code: {
             #     response.status_code}")
            return ""

        count += 1  # Increment count for the next retry

        if count > max_retries:
            print("Maximum retries reached without success.")


def get_postcode_info(postcode):
    url = f'https://api.postcodes.io/postcodes/{postcode}'
    response = requests.get(url)
    if response.status_code == 200:
        data = response.json()
        if data and 'result' in data and data['result']:
            item = data['result']
            postcode_data = {
                'latitude': item['latitude'],
                'longitude': item['longitude'],
                'lsoa': item['codes']['lsoa'],
                'msoa': item['codes']['msoa'],
                'admin_district': item['codes']['admin_district'],
                'police_region': item['codes']['pfa']
            }
            return pd.Series(postcode_data)
    # print(f"Failed to fetch postcode data for {postcode}")
    # print(f"Postfcode failed: {postcode}")
    return pd.Series([None, None, None, None, None, None, None])

def calculate_total_risk_score(row, weightings):
   """
   Calculate total risk score from individual risk scores with dynamic weightings.
   
   Args:
       row: pandas Series containing risk scores
       weightings: dict of risk names and their weights, summing to 1
       
   Returns:
       float: calculated total risk score
   """
   # Copy weightings to avoid modifying original
   adjusted_weights = weightings.copy()
   
   # Remove weights for missing scores
   missing_risks = []
   for risk, weight in weightings.items():
        try:
            if pd.isna(row[risk]) or row[risk] is None:
                missing_risks.append(risk)
        except:
            missing_risks.append(risk)
            continue
   
   # Remove the missing risks from adjusted weights
   for risk in missing_risks:
       adjusted_weights.pop(risk)
   
   # If we have no valid risks, return None or 0
   if not adjusted_weights:
       return None
       
   # Rescale remaining weights to sum to 1
   weight_sum = sum(adjusted_weights.values())
   adjusted_weights = {risk: weight/weight_sum 
                      for risk, weight in adjusted_weights.items()}
   
   # Calculate weighted score
   total_score = 0
   for risk, weight in adjusted_weights.items():
        try:
            total_score += row[risk] * weight
        except:
            continue

       
   return total_score

__all__ = ["run_model"]




def process_all_rows(outputs_df, population_df, data_library):
    # Use the new optimized function directly
    return apply_count_population(
        outputs_df, 
        population_df, 
        6,  # radius_km - adjust this value if needed
        data_library['uk_monthly_police_reporting'],  # adjust these keys to match your data_library
        data_library['ni_monthly_police_reporting']
    )



# At file level with other globals



from scipy.spatial import cKDTree
distance_cache = {}

from concurrent.futures import ProcessPoolExecutor
import multiprocessing as mp
import numpy as np

def process_batch(args):
    batch_coords, population_coords, valid_population_df, monthly_crime_data, uk_df_batch, radius_km_miles = args
    
    # Create local KDTree for this batch
    tree = cKDTree(population_coords)
    batch_results = []
    
    # Query tree for batch
    batch_indices = tree.query_ball_point(batch_coords, radius_km_miles)
    
    # Process each postcode in batch
    for i, indices in enumerate(batch_indices):
        row = uk_df_batch.iloc[i]
        
        within_radius = valid_population_df.iloc[indices]
        sectors = within_radius['sector'].unique().tolist()
        
        crimes_df = pd.DataFrame({'corrected_postcode': [row['corrected_postcode']]})
        
        if len(sectors) > 0:
            observable_crimes_df = monthly_crime_data[monthly_crime_data['Postcode Sector'].isin(sectors)]
            
            for crime in crime_categories:
                if not observable_crimes_df.empty:
                    crime_count = observable_crimes_df[crime].sum() / len(sectors)
                    all_crime_counts = monthly_crime_data[crime]
                    percentile = stats.percentileofscore(all_crime_counts, crime_count)
                    score = assign_score(percentile)
                else:
                    crime_count = 0
                    score = 0
                
                crimes_df[f'{crime} monthly_count'] = [crime_count]
                crimes_df[f'{crime} monthly_rank'] = [score]
        else:
            for crime in crime_categories:
                crimes_df[f'{crime} monthly_count'] = [0]
                crimes_df[f'{crime} monthly_rank'] = [0]
        
        batch_results.append(crimes_df)
    
    return pd.concat(batch_results, ignore_index=True)








# def apply_count_population(df, population_df, radius_km, monthly_crime_data, ni_crime_data):
#     print('Processing population counts...')
    
#     # Print initial counts
#     total_rows = len(population_df)
#     valid_coords = len(population_df.dropna(subset=['latitude', 'longitude']))
#     print(f"\nTotal population rows: {total_rows}")
#     print(f"Rows with valid coordinates: {valid_coords}")
#     print(f"Percentage valid: {(valid_coords/total_rows)*100:.2f}%")
    
#     # Check for zero or invalid values too
#     zeros_mask = (population_df['latitude'] == 0) | (population_df['latitude'] == 0)
#     print(f"Rows with zero coordinates: {zeros_mask.sum()}")
    
#     # Filter out invalid coordinates
#     valid_population_df = population_df.dropna(subset=['latitude', 'longitude'])
#     valid_population_df = valid_population_df[
#         (valid_population_df['latitude'] != 0) & 
#         (valid_population_df['longitude'] != 0)
#     ]
    
#     print(f"Final valid rows for processing: {len(valid_population_df)}")
    
#     # Split dataframes by country to process separately
#     uk_df = df[df['country'].isin(['England', 'Wales'])]
#     ni_df = df[df['country'] == 'Northern Ireland']
    
#     results = []
    
#     if not uk_df.empty:
#         # Process England and Wales using KDTree
#         radius_km_miles = radius_km * 1.60934
        
#         # Use only valid coordinates
#         population_coords = valid_population_df[['latitude', 'longitude']].values
#         query_coords = uk_df[['latitude', 'longitude']].values
        
#         print(f"\nPopulation coords shape: {population_coords.shape}")
#         print(f"Query coords shape: {query_coords.shape}")
        
#         print("Building KD-Tree...")
#         tree = cKDTree(population_coords)
#         print("KD-Tree built, starting queries...")

#         # Process in batches
#         batch_size = 100
#         total_uk = len(uk_df)
        
#         for batch_start in range(0, total_uk, batch_size):
#             batch_end = min(batch_start + batch_size, total_uk)
#             print(f"Processing batch {batch_start+1}-{batch_end} of {total_uk}")
            
#             # Get batch of coordinates
#             batch_coords = query_coords[batch_start:batch_end]
            
#             # Query tree for batch
#             print("Querying KD-Tree...")
#             batch_indices = tree.query_ball_point(batch_coords, radius_km_miles)
#             print("Processing results...")
            
#             # Process each postcode in batch
#             for i, indices in enumerate(batch_indices):
#                 row = uk_df.iloc[batch_start + i]
                
#                 # Use valid_population_df instead of population_df
#                 within_radius = valid_population_df.iloc[indices]
#                 sectors = within_radius['sector'].unique().tolist()
                
#                 if len(sectors) == 0:
#                     print(f"No sectors found for postcode: {row['corrected_postcode']}")
                
#                 crimes_df = pd.DataFrame({'corrected_postcode': [row['corrected_postcode']]})
                
#                 if len(sectors) > 0:
#                     observable_crimes_df = monthly_crime_data[monthly_crime_data['Postcode Sector'].isin(sectors)]
                    
#                     for crime in crime_categories:
#                         if not observable_crimes_df.empty:
#                             crime_count = observable_crimes_df[crime].sum() / len(sectors)
#                             all_crime_counts = monthly_crime_data[crime]
#                             percentile = stats.percentileofscore(all_crime_counts, crime_count)
#                             score = assign_score(percentile)
#                         else:
#                             crime_count = 0
#                             score = 0
                        
#                         crimes_df[f'{crime} monthly_count'] = [crime_count]
#                         crimes_df[f'{crime} monthly_rank'] = [score]
#                 else:
#                     for crime in crime_categories:
#                         crimes_df[f'{crime} monthly_count'] = [0]
#                         crimes_df[f'{crime} monthly_rank'] = [0]
                
#                 results.append(crimes_df)
            
#             print(f"Batch completed. Total results so far: {len(results)}")
    
#     # Process Northern Ireland
#     if not ni_df.empty:
#         total_ni = len(ni_df)
#         for i, (_, row) in enumerate(ni_df.iterrows()):
#             if i % 10 == 0:
#                 print(f"Processing NI row {i+1} of {total_ni}")
                
#             crimes_df = pd.DataFrame({'corrected_postcode': [row['corrected_postcode']]})
#             applicable_row = ni_crime_data[ni_crime_data['admin_district'] == row['admin_district']]
            
#             for crime in crime_categories:
#                 if not applicable_row.empty:
#                     crime_count = applicable_row[crime].iloc[0]
                    
#                     if crime_count is not None:
#                         percentile = stats.percentileofscore(ni_crime_data[crime], crime_count)
#                         score = assign_score(percentile)
#                     else:
#                         crime_count = 0
#                         score = 0
#                 else:
#                     crime_count = 0
#                     score = 0
                
#                 crimes_df[f'{crime}_count'] = [crime_count]
#                 crimes_df[f'{crime} monthly_rank'] = [score]
            
#             results.append(crimes_df)
    
#     print("All processing complete. Combining results...")
    
#     if results:
#         return pd.concat(results, ignore_index=True)
#     else:
#         empty_df = pd.DataFrame(columns=['corrected_postcode'] + 
#                               [f'{crime} monthly_count' for crime in crime_categories] +
#                               [f'{crime} monthly_rank' for crime in crime_categories])
#         return empty_df

def haversine_distance_vectorized(lat1, lon1, lat2, lon2):
    """Calculate haversine distance for vectorized inputs"""
    R = 6371
    lat1, lon1, lat2, lon2 = map(np.radians, [lat1, lon1, lat2, lon2])
    dlat = lat2 - lat1
    dlon = lon2 - lon1
    a = np.sin(dlat/2.0)**2 + np.cos(lat1) * np.cos(lat2) * np.sin(dlon/2.0)**2
    c = 2 * np.arcsin(np.sqrt(a))
    return R * c

# def process_all_rows(outputs_df, population_df, data_library):
#     """Process all rows with explicit dependency passing"""
#     optimal_workers = min(32, multiprocessing.cpu_count() * 2)
    
#     with ThreadPoolExecutor(max_workers=optimal_workers) as executor:
#         temp_monthly_crimes_list = list(executor.map(
#             lambda row_tuple: process_single(row_tuple, population_df, data_library),
#             outputs_df.iterrows()
#         ))
    
#     return pd.concat(temp_monthly_crimes_list, ignore_index=True)


def check_postcode(pc):
    count = 1
    max_retries = 4  # Define a maximum number of retries to prevent an infinite loop

        # Retry logic starts here
    while count <= max_retries:
        url = f'https://api.postcodes.io/postcodes/{pc}/autocomplete/'
        response = requests.get(url)

        if response.status_code == 200:
            data = response.json()
            if data['result']:
                corrected_pc = data['result'][0]
                return corrected_pc
                break  # Success, exit the retry loop
            else:
                print(f'retrying {pc}')
                pc = pc[:count*-1]
        else:
            print(f"Failed to get a successful response, status code: {response.status_code}")

        count += 1  # Increment count for the next retry

        if count > max_retries:
            print("Maximum retries reached without success.")

def get_api_data(postcode):

    postcode = postcode
    corrected_postcode = check_postcode(postcode)

    try:
        url = f'https://api.postcodes.io/postcodes/{corrected_postcode}'
        response = requests.get(url)

        if response.status_code == 200:
            data = response.json()
            result = {'query': postcode, 'result': data['result']}

            return result
        else: 
            return
    except:
        return

def process_postcodes(postcodes):

    with mp.Pool(mp.cpu_count()) as pool:  # Create a pool with number of CPU cores
        results = pool.map(run, postcodes)  # Run 'run' function in parallel for each postcode
    return results
import requests

def fetch_postcode_data(batch):
    url = "https://api.postcodes.io/postcodes"
    payload = {"postcodes": batch}
    response = requests.post(url, json=payload)
    if response.status_code == 200:
        batch_results = response.json()["result"]
        return [result if result["result"] is not None else get_api_data(result["query"]) for result in batch_results]
    return []

def batch_process_postcodes(postcodes, batch_size=100):
    successful_results = []
    with ThreadPoolExecutor(max_workers=10) as executor:
        futures = [executor.submit(fetch_postcode_data, postcodes[i:i + batch_size])
                   for i in range(0, len(postcodes), batch_size)]
        for future in as_completed(futures):
            successful_results.extend(future.result())
    return successful_results




def process_postcode_result(result):
    if result is None: 
        return
    if result["result"] is None:
        return None
    content = result["result"]
    return {
        "postcode": result["query"],
        "corrected_postcode": content['postcode'],
        "lsoa": content["codes"]["lsoa"],
        "msoa": content["codes"]["msoa"],
        "admin_district": content["codes"]["admin_district"],
        "latitude": content["latitude"],
        "longitude": content["longitude"], 
        "police_region": content["codes"]["pfa"]
    }

def process_chunk_monthly(args):
    chunk_df, population_df, monthly_crime_data, ni_monthly_crime_data = args
    return chunk_df.apply(
        lambda row: get_sector_level_crimes(row, population_df, monthly_crime_data, ni_monthly_crime_data),
        axis=1
    )

def process_all_monthly_crime_data(outputs_df, population_df, monthly_crime_data, ni_monthly_crime_data):
    """Process all monthly crime data in parallel with progress bar"""
    num_processes = max(1, cpu_count() - 1)
    total_rows = len(outputs_df)
    
    # For very small datasets, just process directly
    if total_rows < num_processes:
        print(f"\nProcessing {total_rows} rows directly (dataset too small for parallel processing)...")
        results = outputs_df.apply(
            lambda row: get_sector_level_crimes(row, population_df, monthly_crime_data, ni_monthly_crime_data),
            axis=1
        )
        return pd.DataFrame(results)
    
    print(f"\nProcessing {total_rows} rows using {num_processes} processes for monthly crime data...")
    
    # Split DataFrame into chunks, ensuring chunk_size is at least 1
    chunk_size = max(1, len(outputs_df) // num_processes)
    chunks = [outputs_df.iloc[i:i + chunk_size] for i in range(0, len(outputs_df), chunk_size)]
    
    # Prepare arguments for each chunk
    chunk_args = [(chunk, population_df, monthly_crime_data, ni_monthly_crime_data) for chunk in chunks]
    
    # Process chunks in parallel with progress bar
    with Pool(num_processes) as pool:
        results = list(tqdm(
            pool.imap(process_chunk_monthly, chunk_args),
            total=len(chunks),
            desc="Processing monthly crime data"
        ))
    
    print("Combining results...")
    final_results = pd.concat(results)
    print(f"Completed processing {total_rows} rows for monthly crime data")
    return final_results

def get_sector_level_crimes(row, population_df, monthly_crime_data, ni_monthly_crime_data):
    # Initialize empty dictionary to store results
    results = {'corrected_postcode': row['corrected_postcode']}
    
    if row['country'] in ['England', 'Wales']:
        # Clean population DataFrame first
        clean_population_df = population_df.dropna(subset=['latitude', 'longitude'])
        clean_population_df = clean_population_df[~clean_population_df[['latitude', 'longitude']].isin([np.inf, -np.inf]).any(axis=1)]
        
        searcher = SpatialSearch(clean_population_df)
        within_radius = set(searcher.find_within_radius(row['latitude'], row['longitude'], radius_miles=6)['sector'])
        
        crimes_df = monthly_crime_data[monthly_crime_data['Postcode Sector'].isin(within_radius)]
        
        for crime in crime_categories:
            val = crimes_df[crime].values.sum() / len(within_radius)
            percentile = stats.percentileofscore(monthly_crime_data[crime].values, val)
            score = assign_score(percentile)
            
            # Add to results dictionary
            results[f'{crime} monthly_count'] = val
            results[f'{crime} monthly_rank'] = score
    elif row['country'] == 'Northern Ireland':

        try:
            data = ni_monthly_crime_data[ni_monthly_crime_data['admin_district'] == row['admin_district']]

            for crime in crime_categories:
                val = data[crime].values[0]

                if val is None:
                    results[f'{crime} monthly_count'] = None
                    results[f'{crime} monthly_rank'] = None
                else:
                    percentile = stats.percentileofscore(ni_monthly_crime_data[crime].values, val)
                    score = assign_score(percentile)
                    
                    # Add to results dictionary
                    results[f'{crime} monthly_count'] = val
                    results[f'{crime} monthly_rank'] = score
        except:
            for crime in crime_categories:
                results[f'{crime} monthly_count'] = None
                results[f'{crime} monthly_rank'] = None
    else:
        for crime in crime_categories:
            results[f'{crime} monthly_count'] = None
            results[f'{crime} monthly_rank'] = None
    
    return pd.Series(results)



class SpatialSearch:
    def __init__(self, df):
        """
        Initialize the spatial search with a pandas DataFrame.
        
        Args:
            df: pandas DataFrame with 'lat', 'long', and 'sector' columns
        """
        # Store the original dataframe
        self.df = df
        
        # Create numpy array directly from DataFrame columns - very efficient
        self.coords = np.column_stack((df['latitude'].values, df['longitude'].values))
        
        # Create KD-tree for efficient nearest neighbor search
        self.tree = cKDTree(self.coords)
    
    @staticmethod
    def haversine_distance(lat1, lon1, lat2, lon2):
        """
        Calculate the great circle distance between two points 
        on the earth in miles
        """
        R = 3959.87433  # Earth's radius in miles

        lat1, lon1, lat2, lon2 = map(radians, [lat1, lon1, lat2, lon2])
        
        dlat = lat2 - lat1
        dlon = lon2 - lon1
        
        a = sin(dlat/2)**2 + cos(lat1) * cos(lat2) * sin(dlon/2)**2
        c = 2 * atan2(sqrt(a), sqrt(1-a))
        
        return R * c
    
    def find_within_radius(self, lat, lon, radius_miles=6):
        """
        Find all locations within radius_miles of the given lat/lon.
        
        Args:
            lat: Latitude of the center point
            lon: Longitude of the center point
            radius_miles: Radius in miles (default 6)
            
        Returns:
            DataFrame containing matching rows with distances
        """
        # Convert radius to degrees for initial rough filter
        # 1 degree is approximately 69 miles at the equator
        radius_degrees = radius_miles / 69.0
        
        # Use KD-tree to find all points within the bounding box
        indices = self.tree.query_ball_point([lat, lon], radius_degrees)
        
        if not indices:
            return pd.DataFrame(columns=self.df.columns)
        
        # Get the subset of the DataFrame
        results_df = self.df.iloc[indices].copy()
        
        # Calculate exact distances using vectorized operations
        results_df['distance'] = results_df.apply(
            lambda row: self.haversine_distance(lat, lon, row['latitude'], row['longitude']),
            axis=1
        )
        
        # Filter by exact distance
        final_results = results_df[results_df['distance'] <= radius_miles]
        
        # Sort by distance
        return final_results.sort_values('distance')


def run_model(data_library, input_df, chosen_final_score_weightings):
    print(chosen_final_score_weightings)
    # Initial setup and postcode processing
    outputs_df = input_df.copy()
    columns = outputs_df.columns.to_list()
    # print(columns)
    
    # Drop invalid postcodes
    # print(outputs_df)
    postcodes = outputs_df['postcode'].tolist()
    # print(postcodes)
    # postcodes = list(outputs_df)
    
    print('Postcodes to list...')
    # Batch process postcodes
    batch_results = batch_process_postcodes(postcodes)
    print('Got batch results...')
    processed_postcodes = [process_postcode_result(result) for result in batch_results]

    # Create a dictionary for quick lookup
    postcode_data = {item['postcode']: item for item in processed_postcodes}

    postcode_data = {
    postcode: details for postcode, details in postcode_data.items()
    if details.get('latitude') is not None and details.get('longitude') is not None
}

    postcode_df = pd.DataFrame(list(postcode_data.values()))
    # print(postcode_df.head())
    # print(outputs_df.head())

    # Merge with outputs_df
    outputs_df = outputs_df.merge(postcode_df, on='postcode', how='inner')

    # print(outputs_df.head())


    outputs_df['country'] = outputs_df['msoa'].apply(get_country)

    # Get postcode breakdowns
    breakdowns = outputs_df['corrected_postcode'].apply(
        lambda pc: pd.Series(get_postcode_breakdowns(pc))
    )
    outputs_df[['sector', 'district', 'area']] = breakdowns

    # Process Scottish council areas in parallel
    outputs_df['council_area'] = parallel_process_data(
        get_scottish_council_area,
        outputs_df,
        data_library['scottish_postcode_directory']
    )
    

    historic_results = process_all_historic_crime_data(
        outputs_df, 
        data_library['uk_historic_detailed_police_reporting'], 
        data_library['scotland_historic_detailed_police_reporting']
    )

    # # Merge with outputs_df
    outputs_df = pd.merge(outputs_df, historic_results, on='corrected_postcode', how='left')

    results = process_all_monthly_crime_data(outputs_df, data_library['population'], data_library['uk_monthly_police_reporting'], data_library['ni_monthly_police_reporting'])
    outputs_df = pd.merge(outputs_df, results, on='corrected_postcode', how='left')
    # print(outputs_df)

    # Process retail parks
    outputs_df['retail_park_count'], outputs_df['retail_park_rank'] = parallel_process_data(
        get_retail_park_data,
        outputs_df,
        data_library['uk_retail_parks_&_shopping_centres'],
        (1.5,)
    )

    # Process schools
    outputs_df['school_count'], outputs_df['school_rank'] = parallel_process_data(
        get_schools_data,
        outputs_df,
        data_library['schools'],
        (1.5,)
    )

    # Process junctions
    outputs_df['junction_count'], outputs_df['major juntion_rank'] = parallel_process_data(
        get_junctions_data,
        outputs_df,
        data_library['uk_major_junction'],
        (1.5,)
    )

    # Process transport links
    outputs_df['transport_links_count'], outputs_df['national transport links_rank'] = parallel_process_data(
        get_transport_links_data,
        outputs_df,
        data_library['uk_transport_hubs'],
        (1.5,)
    )

    # Process wellbeing data
    outputs_df['wellbeing'], outputs_df['wellbeing_rank'] = parallel_process_data(
        get_wellbeing_data,
        outputs_df,
        (data_library['uk_wellbeing'], data_library['ni_wellbeing'])
    )
    # Process unemployment data
    outputs_df['unemployment'], outputs_df['unemployment_rank'] = parallel_process_data(
        get_unemployment_data,
        outputs_df,
        (
            data_library['uk_unemployment'],
            data_library['ni_unemployment'],
            data_library['scotland_unemployment']
        )
    )

    # Process student data
    outputs_df['student_count'], outputs_df['student population_rank'] = parallel_process_data(
        get_student_data,
        outputs_df,
        (
            data_library['uk_student_population'],
            data_library['ni_student_population'],
            data_library['scotland_student_population']
        )
    )

    # Process population density
    outputs_df['population_density'], outputs_df['population_density_rank'] = parallel_process_data(
        get_population_density_data,
        outputs_df,
        (
            data_library['uk_population_density'],
            data_library['ni_population_density'],
            data_library['scotland_population_density']
        )
    )


    outputs_df['homelessness'], outputs_df['homelessness_rank'] = parallel_process_data(
        get_homelessness_data,
        outputs_df,
        (
            data_library['uk_homelessness'],
            data_library['ni_homelessness'],
            data_library['scotland_homelessness']
        )
    )
    
    # Process pubs data
    outputs_df['pubs_count'], outputs_df['pubs_rank'] = parallel_process_data(
        get_pubs_data,
        outputs_df,
        data_library['pubs']
    )

    # Normalize column names
    outputs_df.columns = [col if col in columns else col.lower() for col in outputs_df.columns]


    # Calculate risk scores in parallel
    risks_list = pd.DataFrame(parallel_process_data(
        assign_risk_scores,
        outputs_df,
        data_sources = None
    ))
    # print(risks_list)
    outputs_df = outputs_df.merge(risks_list, on='corrected_postcode', how='left')

    # Calculate final risk scores
    outputs_df['Risk Output'] = outputs_df.apply(
        lambda row: calculate_total_risk_score(row, chosen_final_score_weightings),
        axis=1
    ).fillna(0)
    
    outputs_df['Risk Score'] = round(outputs_df['Risk Output'], 0)
    outputs_df['Risk Output'] *= 10

    # Save outputs
    datetime_string = datetime.now().strftime('%Y-%m-%d %H-%M-%S')

    user = os.getlogin()
    location = "ASEL"
    if user == "HannahPowell":
        location = "Amberstone Technology Ltd"
    
    output_path = rf"C:\Users\{user}\{location}\Risk - Documents\MODELS\Python Outputs\Raw Output {datetime_string}.csv"
    
    full_output = outputs_df.copy()
    # print(full_output)
    
    # Filter columns for final output
    final_columns = columns 

    for risk, val in chosen_final_score_weightings.items():
        if float(val) > 0:
            print("IN:",risk, val)
            final_columns.append(risk)
        else:
            print("OUT:", risk, val)
    
    final_columns = final_columns + ['Risk Output', 'Risk Score']

    # print(final_columns)
    # print("Available columns:", outputs_df.columns.tolist())
    # print("Looking for:", final_columns)
    outputs_df = outputs_df[final_columns]
    # print(outputs_df)
    
    # Save to CSV and model
    # full_output.to_csv(output_path, index=False)
    # save_model(outputs_df, user, location)
    
    return full_output , outputs_df



def Api(postcode, fsw_val):
        
    # postcodes = pd.DataFrame(['1011 AB' , 'RG2 9UW', 'HP2 4JW', 'N7 6QD', 'SN2 8DA', 'TR7 2BB'],columns=['postcode'])
    # postcodes = ['RG2 9UW', 'HP2 4JW', 'N7 6QD', 'SN2 8DA', 'TR7 2BB']
    postcodes = pd.DataFrame(postcode, columns=['postcode'])
    data_source_library = {
        "uk_monthly_police_reporting": r"Documents/DATA 2/CY Crime/sector_level_crime.csv",  # done
        "schools": r"Documents/DATA 2/Schools/schools_for_script.csv",  # done
        "uk_retail_parks_&_shopping_centres": r"Documents/DATA 2/Retail & Shopping Parks/retail_shopping_park_locations.csv",  # done
        "uk_major_junction": r"Documents/DATA 2/Junctions/Junctions.xlsx",  # done
        "uk_transport_hubs": r"Documents/DATA 2/Transport Hubs/England_Scotland_Wales_transport_hubs.csv",  # done
        "uk_wellbeing": r"Documents/DATA 2/Wellbeing/uk_wellbeing_for_script.csv",  # done
        "uk_unemployment": r"Documents/DATA 2/Unemployment_Student/uk_unemployment_for_script.csv",  # done
        "uk_historic_detailed_police_reporting": r"Documents/DATA 2/Historic Crime/uk_historic_crime_data_for_script.xlsx",  # done
        "scottish_monthly_police_reporting": r"Documents/DATA 2/CY Crime/Police Scotland/Scotland_CY_Crime.csv",  # NOT DONE
        "uk_population_density": r"Documents/DATA 2/Population Density/uk_population_density_for_script.csv",  # done
        "scotland_population_density": r"Documents/DATA 2/Population Density/scotland_population_density_for_script.csv",  # done
        "uk_student_population": r"Documents/DATA 2/Unemployment_Student/uk_student_population_for_script.csv",  # done
        "scotland_student_population": r"Documents/DATA 2/Unemployment_Student/scotland_student_population_for_script.csv",  # done
        "scotland_unemployment": r"Documents/DATA 2/Unemployment_Student/scotland_unemployment_for_script.csv",  # done
        "scotland_historic_detailed_police_reporting": r"Documents/DATA 2/Historic Crime/scotland_historic_crime_data_for_script.csv", # done
        "ni_monthly_police_reporting": r"Documents/DATA 2/CY Crime/ni_monthly_crime_for_script.csv", # done
        "ni_population_density": r"Documents/DATA 2/Population Density/ni_population_density_for_script.csv",  # done
        "ni_wellbeing": r"Documents/DATA 2/Wellbeing/ni_wellbeing_for_script.csv",  # done
        "ni_student_population": r"Documents/DATA 2/Unemployment_Student/ni_student_population_for_script.csv",  # done
        "ni_unemployment": r"Documents/DATA 2/Unemployment_Student/ni_unemployment_for_script.csv",  # done
        "pubs": r"Documents/DATA 2/Pubs/pubs_for_script.csv", # done 
        "scottish_postcode_directory": r"Documents/DATA 2/Historic Crime/scottish_postcode_directory.csv", # done
        "population": r"Documents/DATA 2/Postcodes/ukpostcodes.csv", # done
        "uk_homelessness": r"Documents/DATA 2/Homelessness/uk_homelessness_for_script.csv", # done
        "scotland_homelessness": r"Documents/DATA 2/Homelessness/scotland_homelessness_for_script.csv", # done
        "ni_homelessness": r"Documents/DATA 2/Homelessness/ni_homelessness_for_script.csv" # done
    }


    def load_file(name, path, cache_dir):
        try:
            # full_path = rf"C:\Users\{os.getlogin()}\ASEL\{path}"
            full_path = rf"/Users/saran/{path}"
            # print(full_path)
            # logging.info(f"Attempting to load {name} from {full_path}")

            if not os.path.exists(full_path):
                # logging.error(f"File not found: {full_path}")
                return name, None

            file_extension = os.path.splitext(full_path)[1].lower()
            # logging.info(f"File extension for {name}: {file_extension}")

            if file_extension == '.csv':
                # logging.info(f"Reading CSV file: {name}")
                table = pv.read_csv(full_path)
                df = table.to_pandas()
            elif file_extension == '.xlsx':
                # logging.info(f"Reading Excel file: {name}")
                df = pd.read_excel(full_path)
            elif file_extension == '.ods':
                # logging.info(f"Reading ODS file: {name}")
                df = pd.read_excel(full_path, engine='odf')
            else:
                logging.error(f"Unsupported file format for {
                            name}: {file_extension}")
                return name, None

            if df.empty:
                logging.warning(f"Loaded dataframe for {name} is empty")
            else:
                logging.info(f"Successfully loaded {name}. Shape: {df.shape}")

            return name, df
        except Exception as e:
            logging.error(f"Error loading {name} from {
                        path}: {str(e)}", exc_info=True)
            return name, None


    def load_datasets(data_source_library):
        dataframe_library = {}
        total_files = len(data_source_library)
        cache_dir = os.path.join(os.path.expanduser("~"), ".data_cache")
        os.makedirs(cache_dir, exist_ok=True)

        logging.info(f"Starting to load {total_files} datasets")

        with ThreadPoolExecutor(max_workers=os.cpu_count()) as executor:
            future_to_name = {executor.submit(load_file, name, path, cache_dir): name
                            for name, path in data_source_library.items()}

            for i, future in enumerate(as_completed(future_to_name)):
                name = future_to_name[future]
                try:
                    name, df = future.result()
                    if df is not None and not df.empty:
                        dataframe_library[name] = df
                        logging.info(f"Successfully added {
                                    name} to dataframe_library. Shape: {df.shape}")
                    else:
                        logging.warning(
                            f"Dataset {name} is empty or failed to load")
                except Exception as e:
                    logging.error(f"Error processing {name}: {
                                str(e)}", exc_info=True)

        return dataframe_library


    data_library = load_datasets(data_source_library)

    # print(fsw['Grocery Retail (Default)'])
    # print( postcodes)
    # run_model(data_library, postcodes, fsw['Grocery Retail (Default)'])
    return run_model(data_library, postcodes, fsw[fsw_val])
    